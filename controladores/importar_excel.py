import openpyxl
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)

def leer_excel_preguntas(archivo_path):
    """
    Lee un archivo Excel y extrae el detalle del cuestionario y las preguntas.
    
    Formato esperado del Excel:
    - Fila 1: Encabezados del detalle (Nombre, Tipo, Descripción, Estado)
    - Fila 2: Datos del cuestionario (nombre, tipo, descripción, estado)
    - Fila 3: Encabezados de preguntas
    - Fila 4 en adelante: Preguntas con sus datos
    
    Columnas de preguntas:
        A: Pregunta
        B: Tipo (VF o ALT)
        C: Puntos (1-1000)
        D: Tiempo (segundos, 2-300)
        E: Alternativa 1
        F: Alternativa 2
        G: Alternativa 3 (opcional)
        H: Alternativa 4 (opcional)
        I: Alternativa 5 (opcional)
        J: Alternativa 6 (opcional)
        K: Respuesta Correcta (debe coincidir con una alternativa)
    
    Retorna:
        (detalle, preguntas, errores) donde:
        - detalle es un diccionario con los datos del cuestionario
        - preguntas es una lista de diccionarios
        - errores es una lista de mensajes
    """
    detalle = {}
    preguntas = []
    errores = []
    
    try:
        workbook = load_workbook(archivo_path, data_only=True)
        sheet = workbook.active
        
        # Leer detalle del cuestionario (fila 2)
        detalle_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        if detalle_row and len(detalle_row) >= 4:
            detalle['nombre_formulario'] = str(detalle_row[0]).strip() if detalle_row[0] else ''
            tipo_val = str(detalle_row[1]).strip().upper() if detalle_row[1] else 'I'
            # Normalizar tipo
            if tipo_val.startswith('I') or 'INDIVIDUAL' in tipo_val:
                detalle['tipo_formulario'] = 'I'
            elif tipo_val.startswith('G') or 'GRUPAL' in tipo_val:
                detalle['tipo_formulario'] = 'G'
            else:
                detalle['tipo_formulario'] = 'I'
                errores.append("Tipo de cuestionario no reconocido, se usará 'Individual' por defecto")
            
            detalle['descripcion_formulario'] = str(detalle_row[2]).strip() if detalle_row[2] else ''
            estado_val = str(detalle_row[3]).strip().upper() if detalle_row[3] else 'P'
            # Normalizar estado
            if estado_val.startswith('P') or 'PUBLICO' in estado_val or 'PÚBLICO' in estado_val:
                detalle['estado'] = 'Público'
            elif estado_val.startswith('R') or 'PRIVADO' in estado_val:
                detalle['estado'] = 'Privado'
            else:
                detalle['estado'] = 'Público'
                errores.append("Estado no reconocido, se usará 'Público' por defecto")
        else:
            errores.append("No se encontraron datos del cuestionario en la fila 2")
        
        # Leer preguntas desde la fila 4 (la fila 3 son encabezados)
        for row_num, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            # Saltar filas vacías
            if not row[0] or str(row[0]).strip() == '':
                continue
            
            try:
                pregunta_texto = str(row[0]).strip() if row[0] else ''
                tipo = str(row[1]).strip().upper() if row[1] else 'ALT'
                puntos = row[2] if row[2] is not None else None
                tiempo = row[3] if row[3] is not None else None
                
                # Obtener alternativas (columnas E-J, índices 4-9)
                alternativas = []
                for col_idx in range(4, 10):  # Columnas E a J
                    if col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip():
                        alternativas.append(str(row[col_idx]).strip())
                
                # Respuesta correcta (columna K, índice 10)
                respuesta_correcta = str(row[10]).strip() if len(row) > 10 and row[10] else ''
                
                # Validaciones
                if not pregunta_texto or len(pregunta_texto) < 5:
                    errores.append(f"Fila {row_num}: La pregunta debe tener al menos 5 caracteres")
                    continue
                
                if len(pregunta_texto) > 500:
                    errores.append(f"Fila {row_num}: La pregunta no puede exceder 500 caracteres")
                    continue
                
                if tipo not in ['VF', 'ALT']:
                    errores.append(f"Fila {row_num}: El tipo debe ser 'VF' o 'ALT'")
                    continue
                
                # Validar puntos
                try:
                    puntos = int(puntos) if puntos is not None else 100
                except (ValueError, TypeError):
                    puntos = 100
                
                if puntos <= 0 or puntos > 1000:
                    errores.append(f"Fila {row_num}: Los puntos deben estar entre 1 y 1000")
                    continue
                
                # Validar tiempo
                try:
                    tiempo = int(tiempo) if tiempo is not None else 30
                except (ValueError, TypeError):
                    tiempo = 30
                
                if tiempo < 2 or tiempo > 300:
                    errores.append(f"Fila {row_num}: El tiempo debe estar entre 2 y 300 segundos")
                    continue
                
                # Procesar según el tipo
                if tipo == 'VF':
                    # Verdadero/Falso
                    if respuesta_correcta not in ['Verdadero', 'Falso', 'VERDADERO', 'FALSO', 'V', 'F']:
                        errores.append(f"Fila {row_num}: La respuesta correcta debe ser 'Verdadero' o 'Falso'")
                        continue
                    
                    # Normalizar respuesta
                    if respuesta_correcta.upper() in ['VERDADERO', 'V']:
                        respuesta_correcta = 'Verdadero'
                    else:
                        respuesta_correcta = 'Falso'
                    
                    alternativas = ['Verdadero', 'Falso']
                
                elif tipo == 'ALT':
                    # Alternativas múltiples
                    if len(alternativas) < 2:
                        errores.append(f"Fila {row_num}: Debe tener al menos 2 alternativas")
                        continue
                    
                    if len(alternativas) > 6:
                        errores.append(f"Fila {row_num}: No puede tener más de 6 alternativas")
                        continue
                    
                    # Verificar que la respuesta correcta esté en las alternativas
                    alternativas_lower = [alt.lower() for alt in alternativas]
                    if respuesta_correcta.lower() not in alternativas_lower:
                        errores.append(f"Fila {row_num}: La respuesta correcta '{respuesta_correcta}' no coincide con ninguna alternativa")
                        continue
                    
                    # Encontrar la respuesta correcta exacta (preservar mayúsculas/minúsculas)
                    respuesta_correcta = alternativas[alternativas_lower.index(respuesta_correcta.lower())]
                    
                    # Verificar duplicados
                    if len(alternativas) != len(set(alt.lower() for alt in alternativas)):
                        errores.append(f"Fila {row_num}: No puede tener alternativas duplicadas")
                        continue
                
                # Agregar pregunta válida
                preguntas.append({
                    'nombre_pregunta': pregunta_texto,
                    'tipo_pregunta': tipo,
                    'puntos': puntos,
                    'tiempo': tiempo,
                    'alternativas': alternativas,
                    'respuesta': respuesta_correcta
                })
                
            except Exception as e:
                errores.append(f"Fila {row_num}: Error al procesar - {str(e)}")
                logger.error(f"Error procesando fila {row_num}: {e}")
                continue
        
        workbook.close()
        
        if len(preguntas) == 0 and len(errores) == 0:
            errores.append("No se encontraron preguntas válidas en el archivo Excel")
        
        return detalle, preguntas, errores
        
    except Exception as e:
        logger.error(f"Error al leer archivo Excel: {e}")
        import traceback
        traceback.print_exc()
        errores.append(f"Error al leer el archivo Excel: {str(e)}")
        return {}, [], errores

def crear_plantilla_excel(ruta_archivo):
    """
    Crea un archivo Excel de plantilla para que los docentes puedan llenar con sus preguntas.
    Formato:
    - Fila 1: Detalle del cuestionario (nombre, tipo, descripción, estado)
    - Fila 2: Encabezados de preguntas
    - Fila 3 en adelante: Preguntas
    """
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Cuestionario"
        
        # Estilo
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Fila 1: Detalle del cuestionario
        detalle_headers = [
            'Nombre del Cuestionario',
            'Tipo (I=Individual, G=Grupal)',
            'Descripción',
            'Estado (P=Público, R=Privado)'
        ]
        
        # Ejemplo de detalle
        detalle_ejemplo = [
            'Matemáticas Básicas - Grado 5',
            'I',
            'Cuestionario sobre operaciones básicas',
            'P'
        ]
        
        # Agregar encabezados de detalle
        for col_idx, header in enumerate(detalle_headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Agregar ejemplo de detalle
        for col_idx, valor in enumerate(detalle_ejemplo, start=1):
            cell = sheet.cell(row=2, column=col_idx)
            cell.value = valor
            cell.fill = PatternFill(start_color="e7f3ff", end_color="e7f3ff", fill_type="solid")
        
        # Fila 3: Encabezados de preguntas
        pregunta_headers = [
            'Pregunta',
            'Tipo (VF o ALT)',
            'Puntos (1-1000)',
            'Tiempo (segundos, 2-300)',
            'Alternativa 1',
            'Alternativa 2',
            'Alternativa 3',
            'Alternativa 4',
            'Alternativa 5',
            'Alternativa 6',
            'Respuesta Correcta'
        ]
        
        for col_idx, header in enumerate(pregunta_headers, start=1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ejemplos de preguntas (fila 4 en adelante)
        ejemplos = [
            [
                '¿Cuál es la capital de Perú?',
                'ALT',
                100,
                30,
                'Lima',
                'Cusco',
                'Arequipa',
                'Trujillo',
                '',
                '',
                'Lima'
            ],
            [
                'Python es un lenguaje de programación',
                'VF',
                50,
                20,
                '',
                '',
                '',
                '',
                '',
                '',
                'Verdadero'
            ],
            [
                '¿Cuál es el resultado de 2 + 2?',
                'ALT',
                75,
                25,
                '3',
                '4',
                '5',
                '6',
                '',
                '',
                '4'
            ]
        ]
        
        for row_idx, ejemplo in enumerate(ejemplos, start=4):
            for col_idx, valor in enumerate(ejemplo, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = valor
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 40,  # Nombre / Pregunta
            'B': 20,  # Tipo
            'C': 15,  # Descripción / Puntos
            'D': 15,  # Estado / Tiempo
            'E': 20,  # Alternativa 1
            'F': 20,  # Alternativa 2
            'G': 20,  # Alternativa 3
            'H': 20,  # Alternativa 4
            'I': 20,  # Alternativa 5
            'J': 20,  # Alternativa 6
            'K': 25   # Respuesta Correcta
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Agregar notas/instrucciones
        instruccion_row = 8
        sheet[f'A{instruccion_row}'] = 'INSTRUCCIONES:'
        sheet[f'A{instruccion_row}'].font = Font(bold=True, size=12)
        sheet[f'A{instruccion_row + 1}'] = '1. Fila 1: Encabezados del detalle del cuestionario (NO MODIFICAR)'
        sheet[f'A{instruccion_row + 2}'] = '2. Fila 2: Complete los datos del cuestionario (nombre, tipo, descripción, estado)'
        sheet[f'A{instruccion_row + 3}'] = '3. Fila 3: Encabezados de preguntas (NO MODIFICAR)'
        sheet[f'A{instruccion_row + 4}'] = '4. Desde la fila 4: Agregue sus preguntas'
        sheet[f'A{instruccion_row + 5}'] = '5. Tipo pregunta: VF = Verdadero/Falso, ALT = Alternativas múltiples'
        sheet[f'A{instruccion_row + 6}'] = '6. Tipo cuestionario: I = Individual, G = Grupal'
        sheet[f'A{instruccion_row + 7}'] = '7. Estado: P = Público, R = Privado'
        sheet[f'A{instruccion_row + 8}'] = '8. Para VF, use "Verdadero" o "Falso" en Respuesta Correcta'
        sheet[f'A{instruccion_row + 9}'] = '9. Para ALT, la Respuesta Correcta debe coincidir exactamente con una alternativa'
        sheet[f'A{instruccion_row + 10}'] = '10. Elimine las filas de ejemplo antes de subir su archivo'
        
        workbook.save(ruta_archivo)
        return True
        
    except Exception as e:
        logger.error(f"Error al crear plantilla Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

